import math
from collections import OrderedDict, defaultdict

from flask import Flask, Response, abort, flash, make_response, render_template, request, redirect, send_file, \
    send_from_directory, url_for
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from fractions import Fraction
from flask import Markup


app = Flask(__name__)
config = app.config
app.secret_key = '2lk2ls9a9lqlqk132lk2saa'
app.config['SECRET_KEY'] = '2lk2ls9a9lqlqk132lk2saa'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///project_database.db'
db = SQLAlchemy(app)


# Your model definitions here...
class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False, unique=True)
    form_data_entries = db.relationship('FormDataEntry', backref='project', lazy=True)
    materials = db.relationship('Material', backref='project', lazy=True)
    todo_list_tasks = db.relationship('TodoListTask', backref='project', lazy=True)


class FormLabel(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    category = db.Column(db.String(255))  # Add the category field
    form_data_entries = db.relationship('FormDataEntry', backref='form_label', lazy=True)


class FormDataEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    form_label_id = db.Column(db.Integer, db.ForeignKey('form_label.id'), nullable=False)
    form_quantity = db.Column(db.Integer, nullable=False)


class Material(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)


class TodoListTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    target_quantity = db.Column(db.Integer, nullable=False, default=0)
    current_progress = db.Column(db.Integer, nullable=True, default=0)
    tooltip = db.Column(db.String(255), nullable=True)
    container = db.Column(db.String(50))  # New column for storing container information
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)

    def percentage_complete(self):
        if self.target_quantity == 0:
            return 0
        return math.trunc((self.current_progress / self.target_quantity) * 100)


with app.app_context():
    db.create_all()

global task_data


@app.route('/', methods=['GET', 'POST'])
def home():
    project_id = None  # Initialize project_id

    if request.method == 'POST':
        outlets_quantity = int(request.form.get('outlets'))
        switches_quantity = int(request.form.get('switches'))
        lights_quantity = int(request.form.get('lights'))

        # Create or update FormDataEntry records
        # (You need to implement this based on your model structure)

        # Redirect back to the same page after processing input
        return redirect(url_for('home'))

    return render_template('index.html', project_id=project_id)


@app.route('/bending-calculator', methods=['GET', 'POST'])
def bending_calculator():
    return render_template('bending-calculator.html')

@app.route('/offset', methods=['GET', 'POST'])
def offset():

    picture='offset-labled.png'
    if request.method == 'POST':
        picture='offset-labled-blank.png'
        # Define a dictionary of degree multipliers
        degree_multipliers = {
            10: 6,
            22.5: 2.6,
            30: 2.0,
            45: 1.4,
            60: 1.2
        }
        shrinkage_rates = {
            10: 1 / 16,
            15: 1 / 8,
            22.5: 3 / 16,
            30: 1 / 4,
            45: 3 / 8,
            60: 1 / 2
        }
        offset_angle = float(request.form['offset_angle'])

        # Get the selected degree from the form
        if offset_angle != 22.5:
            offset_angle = int(float(request.form['offset_angle']))
        else:
            offset_angle = offset_angle      

        offset_depth = float(request.form['offset_depth'])  # User-provided offset depth
        shrinkage = offset_depth * math.tan(offset_angle)
        
        if offset_angle in degree_multipliers:
            offset_depth_multiplier = degree_multipliers[offset_angle]
            distance = offset_depth_multiplier * offset_depth
        else:
            # Handle cases where the offset angle is not in the dictionary
            offset_depth = None  # You can define a default value or handle it differently

        if offset_angle in shrinkage_rates:
            shrinkage_rate = shrinkage_rates[offset_angle]
            shrinkage = offset_depth * shrinkage_rate
        else:
            # Handle cases where the offset angle is not in the dictionary
            shrinkage = None  # You can define a default value or handle it differently

        # Round decimal values to 1/8th inch increments and convert to mixed fractions
        offset_depth_rounded= decimal_to_mixed_fraction(round_to_eighth(offset_depth))
        distance_rounded = decimal_to_mixed_fraction(round_to_eighth(distance))
        shrinkage_rounded = decimal_to_mixed_fraction(round_to_eighth(shrinkage))
        degree = f"{offset_angle}°"

        return render_template('offset.html', picture=picture, offset_angle=degree, Fraction=Fraction,
                               offset_depth=offset_depth_rounded, distance_between_bends=distance_rounded, conduit_shrinkage=shrinkage_rounded,
                               math=math)
    offset_depth_mixed = None
    return render_template('offset.html', picture=picture, math=math, offset_depth_mixed=offset_depth_mixed)

def round_to_eighth(number):
    rounded = round(number * 8) / 8

    return rounded

def decimal_to_mixed_fraction(decimal):
    whole_number = int(decimal)
    fraction = Fraction(decimal - whole_number).limit_denominator()
    if whole_number >= 1:
        if int(decimal) == decimal:
            mixed_fraction = Markup(f"{whole_number}\"")
        else:
            mixed_fraction = Markup(f"{whole_number}<sup>{fraction.numerator}</sup>&frasl;<sub>{fraction.denominator}</sub>\"")
    
    else:
        mixed_fraction = Markup(f"<sup>{fraction.numerator}</sup>&frasl;<sub>{fraction.denominator}</sub>\"")
    return mixed_fraction
    
def order(dict):
    final_materials = OrderedDict()
    zero = 0
    # Loop through the items in item_dict and add their quantities from accumulated_materials
    for item, quantity in item_dict.items():
        # Get the quantity from accumulated_materials, if present, else default to 0
        quantity_from_accumulated = dict.get(item, 0)
        # Add the item and quantity to the final_materials dictionary
        if quantity_from_accumulated != zero:
            final_materials[item] = quantity_from_accumulated
    return final_materials


from jobs import joblist


@app.route('/create_project', methods=['GET', 'POST'])
def create_project():
    if request.method == 'POST':
        project_name = request.form.get('project_name')

        try:
            project = Project(name=project_name)
            db.session.add(project)
            db.session.commit()

            return redirect(url_for('add_material', project_id=project.id))
        except IntegrityError:
            db.session.rollback()  # Rollback the transaction

            # Render a page with options for the user
            existing_project = Project.query.filter_by(name=project_name).first()
            return render_template('project_exists.html', project_id=existing_project.id)

    dropdown_options = joblist()
    return render_template('create_project.html', dropdown_options=dropdown_options)


@app.route('/projects/<int:project_id>/edit_form_quantities', methods=['GET', 'POST'])
def edit_form_quantities(project_id):
    # Retrieve the project associated with the provided project_id from the database.
    project = Project.query.get(project_id)
    # Retrieve the form entries related to the project.
    form_entries = project.form_data_entries

    # Retrieve all form labels from the database.
    form_labels = FormLabel.query.all()

    #         material_record.quantity = quantity
    project = Project.query.get(project_id)
    form_labels = FormLabel.query.all()
    form_label_quantities = {}
    if request.method == 'POST':
        form_label_quantities = {}  # Initialize the dictionary to store form label quantities

        for label in form_labels:
            form_label_name = label.name
            form_label_value = request.form.get(form_label_name)

            if form_label_value and form_label_value.strip():
                quantity = int(form_label_value)
            else:
                quantity = 0

            # Find the existing FormDataEntry for the current form label and project.
            form_data_entry = FormDataEntry.query.filter_by(project=project, form_label=label).first()

            if form_data_entry:
                # Update the form_quantity attribute for the existing entry.
                form_data_entry.form_quantity = quantity
            else:
                # If no existing entry is found, create a new one.
                form_data_entry = FormDataEntry(project=project, form_label=label, form_quantity=quantity)
                db.session.add(form_data_entry)

            # Populate the form_label_quantities dictionary.
            form_label_quantities[form_label_name] = quantity

        db.session.commit()

        # Calculate total material requirements based on the updated form_label_quantities.
        total_material_requirements = defaultdict(int)

        for form_label_name, quantity in form_label_quantities.items():
            device_requirements = material_requirements.get(form_label_name, {})
            for material, quantity_per_device in device_requirements.items():
                total_material_requirements[material] += quantity * quantity_per_device

        # Order and add/update material records in the database.
        total_material_requirements = order(total_material_requirements)
        for material, quantity in total_material_requirements.items():
            # Check if a material with the same name and project already exists.
            existing_material = Material.query.filter_by(name=material, project=project).first()

            if existing_material:
                # Update the quantity for the existing material entry.
                existing_material.quantity = quantity
            else:
                # If no existing entry is found, create a new one.
                material_record = Material(name=material, quantity=quantity, project=project)
                db.session.add(material_record)

        db.session.commit()

        form_entries = project.form_data_entries
        device_quantities = {}
        for entry in form_entries:
            device_quantities[entry.form_label.name] = entry.form_quantity
        project, form_entries, device_quantities, duplex, gfci, cutin, surface, controlled, quad_duplex, quad_gfci, quad_cutin, quad_surface, quad_controlled, ff3, ff4, data, data_cutin, HV_dimming, HV_dimming_cutin, LV_cat5, LV_cat5_cutin, HV_switch, HV_cutin, six_in_floor, four_in_floor, wh277v40a, wh208v40a, wh277v30a, singlegang, twogang, outlets, switches, predefined_tasks = new_func(
            project_id)
        # Retrieve the existing tasks associated with the project.
        tasks = TodoListTask.query.filter_by(project_id=project_id).all()

        # Loop through the predefined tasks.
        for task_info in predefined_tasks:
            # Check if a task with the same name and project already exists.
            existing_task = next((task for task in tasks if task.name == task_info['name']), None)

            if existing_task:
                # Update the existing task's attributes.
                existing_task.target_quantity = task_info['target_quantity']
                existing_task.container = task_info['container']
            else:
                # If no existing task is found, create a new one.
                task = TodoListTask(
                    name=task_info['name'],
                    target_quantity=task_info['target_quantity'],
                    project_id=project_id,
                    container=task_info['container']
                )
                db.session.add(task)

        db.session.commit()

        update_or_create_tasks(project_id, predefined_tasks, device_quantities)

        return redirect(url_for('test', project_id=project_id) + '#formQuantities')

    # The GET request handling remains the same.
    return render_template('edit_form_quantities.html', project=project, form_entries=form_entries)

@app.route('/add_material/<project_id>', methods=['GET', 'POST'])
def add_material(project_id):
    project = Project.query.get(project_id)
    form_labels = FormLabel.query.all()
    categories = {}
    for label in form_labels:
        category = label.category
        if category not in categories:
            categories[category] = []
        categories[category].append(label)

    if request.method == 'POST':
        form_label_quantities = {}
        for label in form_labels:
            form_label_name = label.name
            form_label_value = request.form.get(form_label_name)
            if form_label_value and form_label_value.strip():  # Check if value is not empty or only whitespace
                quantity = int(form_label_value)
            else:
                quantity = 0  # Use a default value of 0 if the value is not provided

        db.session.commit()
        form_label_quantities = {}  # Initialize the dictionary outside the loop
        for entry in form_labels:
            form_label_name = entry.name
            form_label_value = request.form.get(form_label_name)
            if form_label_value and form_label_value.strip():
                quantity = int(form_label_value)
            else:
                quantity = 0
            form_label_quantities[form_label_name] = quantity  # Store the quantity in the dictionary            
            form_data_entry = FormDataEntry(project=project, form_label=entry, form_quantity=quantity)
            db.session.add(form_data_entry)

        db.session.commit()

        total_material_requirements = defaultdict(int)
        for form_label_name, quantity in form_label_quantities.items():
            device_requirements = material_requirements.get(form_label_name, {})
            for material, quantity_per_device in device_requirements.items():
                total_material_requirements[material] += quantity * quantity_per_device
        total_material_requirements = order(total_material_requirements)
        for material, quantity in total_material_requirements.items():
            material_record = Material(name=material, quantity=quantity, project=project)
            db.session.add(material_record)
        db.session.commit()

        return redirect(url_for('test', project_id=project_id) + '#projectMaterials')
    return render_template('add_material.html', project=project, categories=categories)


# Route to view all projects
@app.route('/projects')
def view_projects():
    projects = Project.query.all()
    return render_template('projects.html', projects=projects)

def custom_round(value, step=1):
    return round(value / step) * step


# Add the custom filter to the Jinja2 environment
app.jinja_env.filters['custom_round'] = custom_round


@app.route('/download/<int:project_id>')
def download(project_id) -> Response:
    materials = Material.query.filter_by(project_id=project_id).all()
    project = Project.query.filter_by(id=project_id).first()
    # Load the template workbook
    template_filename = 'xlsx/MaterialList.xlsx'
    wb = load_workbook(template_filename)
    ws = wb.active

    # ite data to the merged cell
    for idx, material in enumerate(materials, start=11):
        # Define the range of merged cells for the current row
        merged_range = f"C{idx}:E{idx}"

        # Unmerge the cells temporarily
        ws.unmerge_cells(merged_range)

        # Write the values to individual cells
        ws[f"B{idx}"] = material.quantity
        ws[f"C{idx}"] = material.name

        # Merge the cells back
        ws.merge_cells(merged_range)

    # Save workbook to a temporary file
    ws.unmerge_cells("F8:I8")
    ws["F8"] = project.name
    ws.merge_cells("F8:I8")
    temp_filename = 'temp_material_list.xlsx'
    wb.save(temp_filename)

    # Send the file for download
    return send_file(temp_filename, as_attachment=True, download_name='Material.xlsx')

@app.route('/projects/<int:project_id>/test', methods=['GET', 'POST'])
def test(project_id):
    project = Project.query.get(project_id)
    materials = project.materials
    form_entries = project.form_data_entries
    device_quantities = {}
    for entry in form_entries:
        device_quantities[entry.form_label.name] = entry.form_quantity

    project, form_entries, device_quantities, duplex, gfci, cutin, surface, controlled, quad_duplex, quad_gfci, quad_cutin, quad_surface, quad_controlled, ff3, ff4, data, data_cutin, HV_dimming, HV_dimming_cutin, LV_cat5, LV_cat5_cutin, HV_switch, HV_cutin, six_in_floor, four_in_floor, wh277v40a, wh208v40a, wh277v30a, singlegang, twogang, outlets, switches, predefined_tasks = new_func(
        project_id)

    # ... (existing code)

    # Query the tasks again after creating them
    tasks = TodoListTask.query.filter_by(project_id=project_id).all()
    if not tasks:
        # If no tasks exist, create tasks using predefined data
        for task_info in predefined_tasks:
            task = TodoListTask(
                name=task_info['name'],
                target_quantity=task_info['target_quantity'],
                project_id=project_id,
                container=task_info['container']  # Set the container attribute
            )
            # Set the tooltip based on predefined task information
            # Set the tooltip based on predefined task information and tooltip_keys
            tooltip_parts = [f"{key}: {device_quantities.get(key, 0)}" for key in task_info.get('tooltip_keys', [])]
            task.tooltip = ', '.join(tooltip_parts)

            db.session.add(task)
        db.session.commit()
    task_data = {task.name: task for task in tasks}

    tasks = TodoListTask.query.filter_by(project_id=project_id).all()

    # Create a dictionary to store task data by name
    task_data = {task.name: task for task in tasks}

    container_average = {}
    for task_name, task_info in task_data.items():
        container_name = task_info.container  # Access the container attribute
        if container_name not in container_average:
            container_average[container_name] = []

        # Check if target_quantity is non-zero before performing division
        if task_info.target_quantity != 0:
            container_average[container_name].append((task_info.current_progress / task_info.target_quantity) * 100)
        else:
            container_average[container_name].append(0)

    for container_name, averages in container_average.items():
        if averages:
            container_average[container_name] = sum(averages) / len(averages)
        else:
            container_average[container_name] = 0

    return render_template('test.html', tasks=tasks, task_name=task_name,
                           calculate_progress_percentage=calculate_progress_percentage, project=project,
                           materials=materials, form_entries=form_entries, task_data=task_data, math=math,
                           container_average=container_average, project_id=project_id, singlegang=singlegang,
                           twogang=twogang, outlets=outlets, switches=switches, wh208v40a=wh208v40a,
                           wh277v30a=wh277v30a, wh277v40a=wh277v40a, six_in_floor=six_in_floor,
                           four_in_floor=four_in_floor, LV_cat5=LV_cat5, LV_cat5_cutin=LV_cat5_cutin,
                           HV_switch=HV_switch, HV_dimming_cutin=HV_dimming_cutin, HV_cutin=HV_cutin,
                           HV_dimming=HV_dimming, quad_controlled=quad_controlled, ff3=ff3, ff4=ff4, data=data,
                           data_cutin=data_cutin, quad_duplex=quad_duplex, quad_gfci=quad_gfci, quad_cutin=quad_cutin,
                           quad_surface=quad_surface, duplex=duplex, gfci=gfci, cutin=cutin, surface=surface,
                           controlled=controlled, test=test, device_quantities=device_quantities,
                           fragment_id='formQuantities')


def update_or_create_tasks(project_id, predefined_tasks, device_quantities):
    # Query existing tasks for the specific project
    existing_tasks = TodoListTask.query.filter_by(project_id=project_id).all()

    # Create a dictionary to store existing tasks by name
    existing_task_dict = {task.name: task for task in existing_tasks}

    for task_info in predefined_tasks:
        # Check if the task already exists for this project
        task = existing_task_dict.get(task_info['name'])

        if task:
            # If the task exists, update its attributes
            task.target_quantity = task_info['target_quantity']
            task.container = task_info['container']
        else:
            # If the task doesn't exist, create a new one
            task = TodoListTask(
                name=task_info['name'],
                target_quantity=task_info['target_quantity'],
                project_id=project_id,
                container=task_info['container']
            )
            # Set the tooltip based on predefined task information and tooltip_keys
        tooltip_parts = []
        for key in task_info.get('tooltip_keys', []):
            quantity = device_quantities.get(key, 0)
            tooltip_parts.append(f"{key}: {quantity}")

        task.tooltip = '\n'.join(tooltip_parts)

        db.session.add(task)

    db.session.commit()


def new_func(project_id):
    project = Project.query.get(project_id)
    form_entries = project.form_data_entries
    device_quantities = {entry.form_label.name: entry.form_quantity for entry in form_entries}

    duplex = device_quantities.get('Bracket Box Duplex', 0)
    gfci = device_quantities.get('GFCI', 0)
    cutin = device_quantities.get('Cut-in', 0)
    surface = device_quantities.get('Surface Muonted', 0)
    controlled = device_quantities.get('Controlled', 0)
    quad_duplex = device_quantities.get('Quad Bracket Box', 0)
    quad_gfci = device_quantities.get('Quad GFCI', 0)
    quad_cutin = device_quantities.get('Quad Cut-in', 0)
    quad_surface = device_quantities.get('Quad Surface Mounted', 0)
    quad_controlled = device_quantities.get('Quad Controlled', 0)
    ff3 = device_quantities.get('3wire Furniture Feed', 0)
    ff4 = device_quantities.get('4wire Furniture Feed', 0)
    data = device_quantities.get('Bracket Box Data', 0)
    data_cutin = device_quantities.get('Cut-in Data', 0)
    HV_dimming = device_quantities.get('Line-Voltage Dimming Switch', 0)
    HV_dimming_cutin = device_quantities.get('Line-Voltage Dimming Cutin', 0)
    LV_cat5 = device_quantities.get('LV/Cat5 Switch', 0)
    LV_cat5_cutin = device_quantities.get('LV/Cat5 Cutin', 0)
    HV_switch = device_quantities.get('Line-Voltage Switch', 0)
    HV_cutin = device_quantities.get('Line-Voltage Cut-in', 0)
    six_in_floor = device_quantities.get('6in Floor Device', 0)
    four_in_floor = device_quantities.get('4in Floor Device', 0)
    wh277v40a = device_quantities.get('Single-Pole 277V 40A Instahot', 0)
    wh208v40a = device_quantities.get('2-Pole 208V 40A Instahot', 0)
    wh277v30a = device_quantities.get('Single-Pole 277V 30A Instahot', 0)

    singlegang = duplex + duplex + gfci + controlled + ff3 + ff4 + data + HV_dimming + HV_switch + LV_cat5
    twogang = quad_duplex + quad_gfci + quad_controlled + ff3 + ff4
    outlets = duplex + gfci + controlled + quad_duplex + quad_gfci + quad_controlled + ff3 + ff4 + data
    switches = HV_dimming + HV_switch + LV_cat5
    cutins = HV_dimming_cutin + LV_cat5_cutin + HV_cutin + quad_cutin + cutin + data_cutin + duplex + controlled + quad_duplex + quad_controlled + data
    bracketBoxes = switches + ff3 + ff4 + gfci + quad_gfci,
    sp = "Single-Pole 277V 30A Instahot"
    sixin = "6in Floor Device"
    fourin = "4in Floor Device"
    sp4 = "Single-Pole 277V 40A Instahot"
    tp = "2-Pole 208V 40A Instahot"
    duplexd = "Bracket Box Duplex"
    gfcid = "GFCI"
    cutind = "Cut-in"
    surfaced = "Surface Muonted"
    controlledd = "Controlled"
    quadDupd = "Quad Bracket Box"
    quadGFCId = "Quad GFCI"
    quadCuttind = "Quad Cut-in"
    quadMountedd = "Quad Surface Mounted"
    quadControlledd = "Quad Controlled"
    ff3d = "Furniture Feed"
    ff4d = "4wire Furniture Feed"
    datad = "Bracket Box Data"
    cutd = "Cut-in Data"
    lvds = "Line-Voltage Dimming Switch"
    lvdc = "Line-Voltage Dimming Cutin "
    cats = "LV/Cat5 Switch"
    cat5c = "LV/Cat5 Cutin"
    lvs = "Line-Voltage Switch"
    hvc = "Line-Voltage Cut-in"
    # Create a dictionary to store task data by name
    # Create a dictionary to store task data by name
    predefined_tasks = [

        # duplex                  Bracket Box Duplex                           
        #     gfci                                  gfci = "GFCI"
        #     cutin                                 cutin = "Cut-in"
        #     surface                               surface = "Surface Muonted"
        #     controlled                            controlled = "Controlled"
        #     quad_duplex                           quadDup = "Quad Bracket Box"
        #     quad_gfci                             quadGFCI = "Quad GFCI"
        #     quad_cutin                            quadCuttin = "Quad Cut-in"
        #     quad_surface                          quadMounted = "Quad Surface Mounted"
        #     quad_controlled                       quadControlled = "Quad Controlled"
        #     ff3                                   ff3 = "Furniture Feed"
        #     ff4                                   ff4 = "4wire Furniture Feed"
        #     data                                  data = "Bracket Box Data"
        #     data_cutin                            cutd = "Cut-in Data"
        #     HV_dimming                            lvds = "Line-Voltage Dimming Switch"
        #     HV_dimming_cutin                      lvdc= "Line-Voltage Dimming Cutin "
        #     LV_cat5                               cat5s = "LV/Cat5 Switch"
        #     LV_cat5_cutin                         cat5c = "LV/Cat5 Cutin"
        #     HV_switch                             lvs = "Line-Voltage Switch"
        #     HVCHV_cutin                           hvc = "Line-Voltage Cut-in" "
        #     6in="6in Floor Device                 sp= "Single-Pole 277V 30A Instahot"
        #       sixin="6in Floor Device"
        #   fourin="in Floor Device"
        #   sp4="Single-Pole 277V 40A Instahot"

        ###layout
        {"name": "Typical Office/ General Duty Outlets: ", "target_quantity": duplex,
         "container": "Layout Walls Per Print", "tooltip_keys": [sp, sp4]},
        {"name": "Switches: ", "target_quantity": switches, "container": "Layout Walls Per Print",
         "tooltip_keys": [fourin, sixin]},
        {"name": "Furniture Feeds/ Under Window Outlets: ", "target_quantity": switches,
         "container": "Layout Walls Per Print", "tooltip_keys": [sp4]},
        {"name": "Breakroom GFCI/ Over Counter Outlets:", "target_quantity": switches,
         "container": "Layout Walls Per Print", "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Cut-in Boxes/MP1s:", "target_quantity": switches, "container": "Layout Walls Per Print",
         "tooltip_keys": ["Bracket Box Duplex", "Single-Pole 277V 30A Instahot", cutind]},

        ###Bracket Boxes
        {"name": "Typical Office and General Duty Outlets/Data Outlets:",
         "target_quantity": duplex + controlled + quad_duplex + quad_controlled + data,
         "container": "Screw Up Bracket Boxes", "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Low Voltage/Line Voltage Switches:", "target_quantity": switches,
         "container": "Screw Up Bracket Boxes", "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Furniture Feeds:", "target_quantity": ff3 + ff4, "container": "Screw Up Bracket Boxes",
         "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Breakroom Outlets/ Over Counter Outlets:", "target_quantity": gfci + quad_gfci,
         "container": "Screw Up Bracket Boxes", "tooltip_keys": ["Bracket Box Duplex"]},

        ###Cutin Boxes
        {"name": "Duplex/Quad Outlet:", "target_quantity": cutin + quad_cutin, "container": "Cut in all Cutins/MP1s",
         "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Data/Low Voltage Switch Cutins:", "target_quantity": data_cutin + LV_cat5_cutin,
         "container": "Cut in all Cutins/MP1s", "tooltip_keys": ["Bracket Box Duplex"]},

        ###TV Boxes
        {"name": "Mount TV Boxes:", "target_quantity": data_cutin + LV_cat5_cutin, "container": "Install TV Boxes",
         "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Drop MC:", "target_quantity": data_cutin + LV_cat5_cutin, "container": "Install TV Boxes",
         "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Drop Flex and String:", "target_quantity": data_cutin + LV_cat5_cutin,
         "container": "Install TV Boxes", "tooltip_keys": ["Bracket Box Duplex"]},

        ###Wall Penetrations
        {"name": "Penetrate Floor for All Floor Devices:", "target_quantity": data_cutin + LV_cat5_cutin,
         "container": "Wall Penetrations", "tooltip_keys": ["Bracket Box Duplex"]},

        ###Loop out all MC
        {"name": "Typical Office and General Duty Outlets/Data Outlets:",
         "target_quantity": duplex + controlled + quad_duplex + quad_controlled + data,
         "container": "Loop MC To all Box Locations", "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Drop MC to Cut-in Boxes:", "target_quantity": cutin + quad_cutin,
         "container": "Loop MC To all Box Locations", "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Dedicated Breakroom Outlets/ Overcounter Outlets:", "target_quantity": cutin + quad_cutin,
         "container": "Loop MC To all Box Locations", "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Furniture Feeds/ Undercounter Outlets:", "target_quantity": cutin + quad_cutin,
         "container": "Loop MC To all Box Locations", "tooltip_keys": ["Bracket Box Duplex"]},
        {"name": "Pull HRs to HR box in ceiling:", "target_quantity": cutin + quad_cutin,
         "container": "Loop MC To all Box Locations", "tooltip_keys": ["Bracket Box Duplex"]}

    ]

    return project, form_entries, device_quantities, duplex, gfci, cutin, surface, controlled, quad_duplex, quad_gfci, quad_cutin, quad_surface, quad_controlled, ff3, ff4, data, data_cutin, HV_dimming, HV_dimming_cutin, LV_cat5, LV_cat5_cutin, HV_switch, HV_cutin, six_in_floor, four_in_floor, wh277v40a, wh208v40a, wh277v30a, singlegang, twogang, outlets, switches, predefined_tasks


def calculate_progress_percentage(current_progress, target_quantity):
    if target_quantity != 0:
        return math.trunc((current_progress / target_quantity) * 100)
    else:
        return 0


@app.route('/projects/<int:project_id>/delete', methods=['POST'])
def delete_project(project_id):
    if request.method == 'POST':
        project = Project.query.get_or_404(project_id)

        # Delete all form data entries associated with the project
        FormDataEntry.query.filter_by(project_id=project_id).delete()

        # Delete all materials associated with the project
        Material.query.filter_by(project_id=project_id).delete()

        # Delete all todo list tasks associated with the project
        TodoListTask.query.filter_by(project_id=project_id).delete()

        # Commit the changes to the database
        Project.query.filter_by(id=project.id).delete()
        db.session.commit()

        # Optionally, you can redirect to a page or return a response
        return redirect(url_for('view_projects'))
    else:
        return abort(405)  # Method Not Allowed


@app.template_filter('format_as_integer')
def format_as_integer(value):
    return "{:.0f}".format(value)


@app.route('/projects/<int:project_id>/update_progress', methods=['POST'])
def update_progress(project_id):
    tasks = TodoListTask.query.filter_by(project_id=project_id).all()
    task_data = {task.name: task for task in tasks}
    device_quantities = new_func(project_id)
    predefined_tasks = new_func(project_id)
    # Calculate average progress percentages for each container

    for task_name, task in task_data.items():
        progress_key = f"progress_{task_name}"
        if progress_key in request.form:
            new_progress = int(request.form[progress_key])

            if task.target_quantity != 0:  # Check for division by zero
                task.current_progress = math.trunc(new_progress)
                task.rounded_progress_percentage = (new_progress / task.target_quantity) * 100

                db.session.commit()
            else:
                flash("Target quantity cannot be zero. Progress percentage calculation skipped.", "error")

            # Update the database with the new current_progress value
            db.session.commit()  # Commit the changes to the database

    return redirect(url_for('test', project_id=project_id) + '#todo_list')


@app.route('/projects/<int:project_id>/submit_progress', methods=['POST'])
def submit_progress(project_id):
    project = Project.query.get(project_id)
    form_entries = project.form_data_entries

    for entry in form_entries:
        input_field_name = f'progressInput_{entry.id}'
        user_input = int(request.form.get(input_field_name))

        # Update the user_progress value in the database
        entry.user_progress = user_input
        db.session.commit()

    return redirect(url_for('table', project_id=project_id))  # Redirect to the updated table page


@app.route('/update_task_progress/<int:task_id>', methods=['POST'])
def update_task_progress(task_id):
    task = TodoListTask.query.get(task_id)
    if task:
        progress = int(request.form.get('progress'))
        task.current_progress = progress
        task.is_complete = (progress >= task.target_quantity)
        db.session.commit()
    return redirect(url_for('project_detail', project_id=task.project_id))


@app.route('/manifest.json')
def manifest():
    return send_from_directory('static', 'manifest.json')


@app.route('/sw.js')
def service_worker():
    response = make_response(send_from_directory('static', 'sw.js'))
    response.headers['Cache-Control'] = 'no-cache'
    return response


item_dict = OrderedDict([
    ('Single Gang Mud Ring', 0),
    ('Two Gang Mud Ring', 0),
    ('4-Square Bracket Box', 0),
    ('Deep 4-Square Bracket Box', 0),
    ('4-Square Box', 0),
    ('Deep 4-Square Box', 0),
    ('4-Square Cover', 0),
    ('Cut-In Box', 0),
    ('Drywall Clamps', 0),
    ('NVent Caddy Mounting Slider Bracket', 0),
    ('12/2 LV MC', 0),
    ('12/3 LV MC', 0),
    ('12/2 HV MC', 0),
    ('12/3 HV MC', 0),
    ('10/2 HV MC', 0),
    ('8/3 LV MC', 0),
    ('18/2 LV Dimmer Cable', 0),
    ('Single Barrel MC Connector', 0),
    ('Double Barrel MC Connector', 0),
    ('Ground Stinger', 0),
    ('Tek Screws', 0),
    ('1/2" Panhead Selftapper', 0),
    ('Mac-2 Straps', 0),
    ('Red Heads', 0),
    ('Red/Yellow Wire Nuts', 0),
    ('Blue/Orange Wire Nuts', 0),
    ('Big Blue Wire Nuts', 0),
    ('Jet Line', 0),
    ('3/4” Snap-In Bushings', 0),
    ('1" Snap-In Bushings', 0),
    ('Single Gang LV1s', 0),
    ('10ft Pieces of 2" EMT **May need adjusting per floor device specs**', 0),
    ('2" EMT Coupling **May need adjusting per floor device specs**', 0),
    ('2" EMT to Flex Change Over **May need adjusting per floor device specs**', 0),
    ('2" 90° Elbow **May need adjusting per floor device specs**', 0),
    ('10ft Pieces of 2" Flex **May need adjusting per floor device specs**', 0),
    ('2" Insulating Push On Conduit Bushing **May need adjusting per floor device specs**', 0),
    ('2" Min Strap **May need adjusting per floor device specs**', 0),
    ('Tube of Fire Cock **May need adjusting per floor device specs**', 0),
    ('1/4” Toggle Bolts', 0),
    ('1/2" EMT', 0),
    ('1/2" EMT Connectors', 0),
    ('1/2" One Hole Strap', 0),
    ('#12 THHN Black Wire', 0),
    ('#12 THHN White Wire', 0),
    ('#12 THHN Green Wire', 0),
    ('Standard Outlet', 0),
    ('Decora Outlet', 0),
    ('GFCI Outlet', 0),
    ('Half-Duplex Controlled Standard Outlet', 0),
    ('Half-Duplex Controlled Decora Outlet', 0),
    ('Full-Duplex Controlled Standard Outlet', 0),
    ('Full-Duplex Controlled Decora Outlet', 0),
    ('Single-Pole Motor Rated 30A Switch', 0),
    ('Two-Pole Motor Rated 40A Switch', 0),
    ('Single Gang Standard Outlet Cover Plate', 0),
    ('Single Gang Decora Outlet Cover Plate', 0),
    ('Two Gang Standard Outlet Cover Plate', 0),
    ('Two Gang Decora Outlet Cover Plate', 0),
    ('4-Square Industrial Switch Cover Plate', 0),
    ('Single Gang Standard Industrial Cover Plug Plate', 0),
    ('Single Gang Decora Industrial Cover Plug Plate', 0),
    ('2 Gang Standard Industrial Cover Plug Plate', 0),
    ('2 Gang Decora Industrial Cover Plug Plate', 0),
    ('3/4" MC Connector', 0),
    ('Two Gang Stainless Steel Blank Plate', 0),
    ('90 Degree 1/2" Flex Connector', 0),
    ('4" Floor Device ***Per Print***', 0),
    ('6" Floor Device ***Per Print***', 0),
    ('Floor Device to Flex Converter ***Per Print***', 0),
    ('Multifunction Clip w/ nut or bolt', 0),
    ('KX Straps', 0),
    ('Ceiling Wires', 0),
])

material_requirements = {
    'Bracket Box Duplex': {
        '4-Square Bracket Box': 1,
        'Single Gang Mud Ring': 1,
        'Ground Stinger': 1,
        'Tek Screws': 10,
        'Mac-2 Straps': 4,
        'Red/Yellow Wire Nuts': 1,
        'Red Heads': 2,
        'Double Barrel MC Connector': 1,
        '12/2 LV MC': 30,
        'KX Straps': 2,
        'Typical Outlet': 1,
        'Single Gang Typical Outlet Cover Plate': 1
    },

    'GFCI': {
        '4-Square Bracket Box': 1,
        'Single Gang Mud Ring': 1,
        'Ground Stinger': 1,
        'Tek Screws': 10,
        'Mac-2 Straps': 4,
        'Red/Yellow Wire Nuts': 1,
        'Red Heads': 2,
        'Single Barrel MC Connector': 1,
        'NVent Caddy Mounting Slider Bracket': 1,
        '12/2 LV MC': 30,
        'KX Straps': 2,
        'GFCI Outlet': 1
    },

    'Cut-In': {

        '4-Square Bracket Box': 1,
        'Ground Stinger': 1,
        'Tek Screws': 10,
        'Mac-2 Straps': 4,
        'Red/Yellow Wire Nuts': 1,
        'Red Heads': 2,
        'Single Barrel MC Connector': 1,
        'NVent Caddy Mounting Slider Bracket': 0,
        '12/2 LV MC': 30,
        'KX Straps': 2,
        'Typical Outlet': 1,
        'Single Gang Typical Outlet Cover Plate': 1,
        'Cut-In Box': 1,
        'Drywall Clamps': 1
    },

    'Surface Mounted': {

        '4-Square Box': 2,
        'Single Gang Typical Industrial Cover Plug Plate': 1,
        '4-Square Cover': 1,
        'Ground Stinger': 2,
        '1/4” Toggle Bolts': 6,
        '1/2" EMT': 10,
        '1/2" EMT Connectors': 2,
        '1/2" One Hole Strap': 2,
        'Double Barrel MC Connector': 2,
        'Red Heads': 2,
        'Red/Yellow Wire Nuts': 6,
        'Mac-2 Straps': 1,
        'KX Straps': 2,
        '#12 THHN Black Wire': 15,
        '#12 THHN White Wire': 15,
        '#12 THHN Green Wire': 15,
        'Typical Outlet': 1,
    },

    'Controlled': {

        'Deep 4-Square Bracket Box': 2,
        'Single Gang Mud Ring': 1,
        'Ground Stinger': 2,
        'Tek Screws': 10,
        'Mac-2 Straps': 3,
        'Red/Yellow Wire Nuts': 5,
        'Red Heads': 4,
        'Double Barrel MC Connector': 1,
        'Single Barrel MC Connector': 2,
        'NVent Caddy Mounting Slider Bracket': 2,
        '12/2 LV MC': 20,
        '12/3 LV MC': 10,
        'Half-Duplex Controlled Typical Outlet': 1,
        'Single Gang Typical Outlet Cover Plate': 1,
    },

    # ----------------------------------------------------------------------

    'Quad Bracket Box': {

        '4-Square Bracket Box': 1,
        'Single Gang Mud Ring': 1,
        'Ground Stinger': 1,
        'Tek Screws': 10,
        'Mac-2 Straps': 4,
        'Red/Yellow Wire Nuts': 1,
        'Red Heads': 2,
        'Double Barrel MC Connector': 1,
        '12/2 LV MC': 30,
        'KX Straps': 2,
        'Typical Outlet': 2,
        'Two Gang Typical Outlet Cover Plate': 1,
    },

    # def calc_quad_decora: {
    #     global accumulated_materials, color_mode
    #     quantity = int(quantity)  # Convert the quantity to an integer
    #     if quantity != 0:
    #         materials = {
    #           '4-Square Bracket Box': 1,
    #           'Single Gang Mud Ring': 1,
    #           'Ground Stinger': 1,
    #           'Tek Screws': 10,
    #           'Mac-2 Straps': 4,
    #           'Red/Yellow Wire Nuts': 1,
    #           'Red Heads': 2,
    #           'Double Barrel MC Connector': 1,
    #           'NVent Caddy Mounting Slider Bracket': 0,
    #           '12/2 LV MC': 30,
    #           'KX Straps': 2,
    #           'Decora Outlet': 2,
    #           'Two Gang Decora Outlet Cover Plate': 1,
    #         }
    #         # Update accumulated_materials with the quantities from materials
    #         for item, quantity in materials.items():
    #             accumulated_materials[item] = accumulated_materials.get(item, 0) + quantity

    'Quad GFCI': {

        '4-Square Bracket Box': 1,
        'Single Gang Mud Ring': 1,
        'Ground Stinger': 1,
        'Tek Screws': 10,
        'Mac-2 Straps': 4,
        'Red/Yellow Wire Nuts': 1,
        'Red Heads': 2,
        'Double Barrel MC Connector': 1,
        'NVent Caddy Mounting Slider Bracket': 1,
        '12/2 LV MC': 30,
        'KX Straps': 2,
        'GFCI Outlet': 2,
        'Two Gang Typical Outlet Cover Plate': 1,
    },

    'Quad Cut-in': {

        'Ground Stinger': 1,
        'Mac-2 Straps': 4,
        'Red/Yellow Wire Nuts': 1,
        'Red Heads': 2,
        'Double Barrel MC Connector': 1,
        'Single Barrel MC Connector': 2,
        '12/2 LV MC': 30,
        'KX Straps': 2,
        'Typical Outlet': 2,
        'Single Gang Typical Outlet Cover Plate': 1,
        'Cut-In Box': 2,
        'Drywall Clamps': 1
    },

    'Quad Surface Mounted': {

        '4-Square Box': 2,
        '2 Gang Typical Industrial Cover Plug Plate': 1,
        '4-Square Cover': 1,
        'Ground Stinger': 2,
        '1/4” Toggle Bolts': 6,
        '1/2" EMT': 10,
        '1/2" EMT Connectors': 2,
        '1/2" One Hole Strap': 2,
        'Double Barrel MC Connector': 2,
        'Red Heads': 2,
        'Red/Yellow Wire Nuts': 6,
        'KX Straps': 2,
        '#12 THHN Black Wire': 15,
        '#12 THHN White Wire': 15,
        '#12 THHN Green Wire': 15,
        'Typical Outlet': 2,
    },

    'Quad Controlled': {

        'Deep 4-Square Bracket Box': 2,
        'Two Gang Mud Ring': 1,
        'Ground Stinger': 2,
        'Tek Screws': 10,
        'Mac-2 Straps': 6,
        'Red/Yellow Wire Nuts': 5,
        'Red Heads': 4,
        'Double Barrel MC Connector': 1,
        'Single Barrel MC Connector': 2,
        'NVent Caddy Mounting Slider Bracket': 3,
        '12/2 LV MC': 20,
        '12/3 LV MC': 10,
        'Full-Duplex Controlled Typical Outlet': 1,
        'Single Gang Typical Outlet Cover Plate': 1,
    },

    # -------------------------------------------------------------------------

    '3wire Furniture Feed': {

        '4-Square Bracket Box': 1,
        'Two Gang Mud Ring': 1,
        'Ground Stinger': 1,
        'Tek Screws': 10,
        'Mac-2 Straps': 4,
        'Red/Yellow Wire Nuts': 8,
        'Red Heads': 1,
        'Single Barrel MC Connector': 1,
        '12/3 LV MC': 30,
        'Two Gang Stainless Steel Blank Plate': 1,
        '90 Degree 1/2" Flex Connector': 1,
    },

    '4wire Furniture Feed': {

        '4-Square Bracket Box': 1,
        'Two Gang Mud Ring': 1,
        'Ground Stinger': 1,
        'Tek Screws': 10,
        'Mac-2 Straps': 4,
        'Red/Yellow Wire Nuts': 8,
        'Red Heads': 4,
        'Double Barrel MC Connector': 2,
        '12/3 LV MC': 30,
        'Two Gang Stainless Steel Blank Plate': 1,
        '90 Degree 1/2" Flex Connector': 1,
        'NVent Caddy Mounting Slider Bracket': 1,
    },

    'Bracket Box Data': {

        '4-Square Bracket Box': 1,
        'Single Gang Mud Ring': 1,
        'Tek Screws': 4,
        'Jet Line': 10,
        '3/4” Snap-In Bushings': 1,
        '1" Snap-In Bushings': 1,
    },

    'Cut-in Data': {

        'LV1s': 1,
        'Jet Line': 10,
    },

    'Line-Voltage Dimming Switch': {

        '12/3 HV MC': 10,
        'Ground Stinger': 1,
        'Single Barrel MC Connector': 1,
        'Deep 4-Square Bracket Box': 1,
        'Double Barrel MC Connector': 2,
        'Red Heads': 5,
        'Red/Yellow Wire Nuts': 8,
        '18/2 LV Dimmer Cable': 15,
        'Blue/Orange Wire Nuts': 2,
        '4-Square Cover': 1,
        'Cut-In Box': 1,
        'Drywall Clamps': 1
    },

    'Line-Voltage Dimming': {

        '4-Square Bracket Box': 1,
        'Single Gang Mud Ring': 1,
        '12/3 HV MC': 10,
        'Mac-2 Straps': 3,
        'Ground Stinger': 1,
        'Single Barrel MC Connector': 1,
        'Deep 4-Square Bracket Box': 1,
        'Double Barrel MC Connector': 2,
        'Red Heads': 1,
        'Tek Screws': 6,
        'Red/Yellow Wire Nuts': 3,
        '18/2 LV Dimmer Cable': 15,
        'Blue/Orange Wire Nuts': 2,
    },

    'LV/Cat5 Switch': {

        '4-Square Bracket Box': 1,
        'Single Gang Mud Ring': 1,
        'Jet Line': 10,
        '3/4” Snap-In Bushings': 1,
        'Deep 4-Square Bracket Box': 1,
        'Double Barrel MC Connector': 2,
        'Tek Screws': 6,
        'KX Straps': 4,
        '4-Square Cover': 1,
        'Ground Stinger': 1,
        'Red/Yellow Wire Nuts': 1,
        '12/2 HV MC': 15,
        'Ceiling Wires': 1,
    },

    'Lv/Cat5 Cut-in': {

        'Jet Line': 10,
        'LV1s': 1,
        'Deep 4-Square Bracket Box': 1,
        'Double Barrel MC Connector': 2,
        'Tek Screws': 3,
        'KX Straps': 4,
        '4-Square Cover': 1,
        'Ground Stinger': 1,
        'Red/Yellow Wire Nuts': 5,
        '12/2 HV MC': 15,
        'Ceiling Wires': 1,
    },

    'Line-Voltage Switch': {

        '4-Square Bracket Box': 1,
        'Single Gang Mud Ring': 1,
        '12/3 HV MC': 10,
        'Mac-2 Straps': 3,
        'Ground Stinger': 2,
        'Deep 4-Square Bracket Box': 1,
        'Double Barrel MC Connector': 2,
        'Single Barrel MC Connector': 1,
        'Red Heads': 1,
        'Tek Screws': 6,
        'Red/Yellow Wire Nuts': 3,

    },

    'Line-Voltage Cut-in': {

        '12/3 HV MC': 10,
        'Ground Stinger': 2,
        'Deep 4-Square Bracket Box': 1,
        '4-Square Cover': 1,
        'Double Barrel MC Connector': 2,
        'Single Barrel MC Connector': 1,
        'Red Heads': 5,
        'Red/Yellow Wire Nuts': 8,
        'Cut-In Box': 1,
        'Drywall Clamps': 1
    },

    '6in Floor Device': {

        '10ft Pices of 2" EMT **May need adjusting per floor device specs**': 1,
        '2" EMT Coupling **May need adjusting per floor device specs**': 1,
        '2" EMT to Flex Change Over **May need adjusting per floor device specs**': 2,
        '2" 90° Elbow **May need adjusting per floor device specs**': 1,
        '10ft Pieces of 2" Flex **May need adjusting per floor device specs**': 1,
        '2" Insulating Push On Conduit Bushing **May need adjusting per floor device specs**': 1,
        '2" Min Strap **May need adjusting per floor device specs**': 5,
        'Tube of Fire Cock **May need adjusting per floor device specs**': 1,
        'Jet Line': 30,
        'Multifuction Clip w/ nut or bolt': 2,
        '6" Floor Device ***Per Print***': 1,
        'Floor Device to Flex Converter ***Per Print***': 10,
        '1/2" Panhead Selftapper': 5,
        'Mac-2 Straps': 3,
        'Red/Yellow Wire Nuts': 8,
        'Red Heads': 2,
        'Single Barrel MC Connector': 2,
        '12/2 LV MC': 30,
        '4-Square Bracket Box': 1,
        '4-Square Cover': 1,
        'KX Straps': 5,

    },

    '4in Floor Device': {

        'Tube of Fire Cock **May need adjusting per floor device specs**': 1,
        '4" Floor Device ***Per Print***': 1,
        'Mac-2 Straps': 3,
        'Red/Yellow Wire Nuts': 8,
        'Red Heads': 2,
        'Single Barrel MC Connector': 2,
        '12/2 LV MC': 30,
        '4-Square Bracket Box': 1,
        '4-Square Cover': 1,
        'KX Straps': 5,

    },

    '2-pole 208V 40A Instahot': {

        '3/4" MC Connector': 4,
        'Deep 4-Square Box': 1,
        '4-Square Industrial Switch Cover Plate': 1,
        '1/2" One Hole Strap': 3,
        'Big Blue Wire Nuts': 4,
        'Two-Pole Motor Rated 40A Switch': 1,
        '8/3 LV MC': 30,
    },

    'Single-Pole 277V 30A Instahot': {

        '3/4" MC Connector': 4,
        'Deep 4-Square Box': 1,
        '4-Square Industrial Switch Cover Plate': 1,
        '1/2" One Hole Strap': 3,
        'Big Blue Wire Nuts': 4,
        'Single-Pole Motor Rated 30A Switch': 1,
        '10/2 HV MC': 30,
    }

}

if __name__ == '__main__':
    app.run(debug=True)
